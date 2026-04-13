from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.config import OUTPUT_DIR

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
DIRECT_FILL = PatternFill("solid", fgColor="E2F0D9")
ADJACENT_FILL = PatternFill("solid", fgColor="FFF2CC")
EXCLUDE_FILL = PatternFill("solid", fgColor="F4CCCC")
INPUT_FILL = PatternFill("solid", fgColor="EAF2F8")
LINK_FONT = Font(color="0563C1", underline="single")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_FONT = Font(color="000000", bold=False)

FINAL_DECISION_OPTIONS = ["Direct", "Adjacent", "Exclude", "Hold"]
ACTION_STATUS_OPTIONS = ["new", "reviewed", "need_spec_check", "need_biz_check", "closed"]


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ",".join(str(x) for x in value if x not in (None, ""))
    return str(value)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _style_body(ws) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _set_column_widths(ws, max_width: int = 42) -> None:
    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        adjusted = min(max(max_len + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _apply_label_fill(ws, label_col_name: str = "label") -> None:
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    label_col_idx = header_map.get(label_col_name)
    if not label_col_idx:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=label_col_idx)
        value = str(cell.value or "").strip()

        if value == "Direct":
            cell.fill = DIRECT_FILL
        elif value == "Adjacent":
            cell.fill = ADJACENT_FILL
        elif value == "Exclude":
            cell.fill = EXCLUDE_FILL


def _apply_url_links(ws, header_names: list[str]) -> None:
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for header_name in header_names:
        col_idx = header_map.get(header_name)
        if not col_idx:
            continue

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = str(cell.value or "").strip()
            if value.startswith("http://") or value.startswith("https://"):
                cell.hyperlink = value
                cell.font = LINK_FONT


def _highlight_input_columns(ws, header_names: list[str]) -> None:
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for header_name in header_names:
        col_idx = header_map.get(header_name)
        if not col_idx:
            continue

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = INPUT_FILL


def _add_dropdown_validation(ws, header_name: str, options: list[str]) -> None:
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    col_idx = header_map.get(header_name)
    if not col_idx:
        return

    last_row = max(ws.max_row, 1000)
    formula = '"' + ",".join(options) + '"'
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
    )
    validation.error = "허용된 값만 입력 가능"
    validation.errorTitle = "입력값 오류"
    validation.prompt = "목록에서 값을 선택"
    validation.promptTitle = header_name
    ws.add_data_validation(validation)
    validation.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row}")


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    _style_header(ws)
    _style_body(ws)
    _set_column_widths(ws)


def build_candidates_rows(notices: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "review_scope",
        "label",
        "score",
        "record_id",
        "source_kind",
        "source_type",
        "title",
        "pub_org",
        "demand_org",
        "close_dt",
        "query_keywords",
        "query_tracks",
        "reasons",
        "linked_bid_count",
        "linked_bid_ntce_nos",
        "bidNtceNoList",
        "bid_detail_url",
        "detail_url_missing_reason",
        "swBizObjYn",
        "purchsObjPrdctList",
        "prdctDtlList",
        "bid_attachment_count",
        "prespec_attachment_count",
        "attachment_job_count",
        "attachment_detected_from",
        "attachment_selected",
        "attachment_missing_reason",
    ]

    rows: list[list[Any]] = []
    for item in notices:
        linked_bid_nos = item.get("_linked_bid_ntce_nos", [])
        rows.append(
            [
                "Y" if item.get("_label") in {"Direct", "Adjacent"} else "N",
                _to_text(item.get("_label")),
                item.get("_score", 0),
                _to_text(item.get("_record_id")),
                _to_text(item.get("_source_kind")),
                _to_text(item.get("_source_type")),
                _to_text(item.get("_title")),
                _to_text(item.get("_pub_org")),
                _to_text(item.get("_demand_org")),
                _to_text(item.get("_close_dt")),
                _to_text(item.get("_query_keywords")),
                _to_text(item.get("_query_tracks")),
                _to_text(item.get("_reasons")),
                len(linked_bid_nos),
                _to_text(linked_bid_nos),
                _to_text(item.get("bidNtceNoList")),
                _to_text(item.get("_bid_detail_url")),
                _to_text(item.get("_detail_url_missing_reason")),
                _to_text(item.get("swBizObjYn")),
                _to_text(item.get("purchsObjPrdctList")),
                _to_text(item.get("prdctDtlList")),
                item.get("_bid_attachment_count", 0),
                item.get("_prespec_attachment_count", 0),
                item.get("_attachment_job_count", 0),
                _to_text(item.get("_attachment_detected_from")),
                _to_text(item.get("_attachment_selected")),
                _to_text(item.get("_attachment_missing_reason")),
            ]
        )

    return headers, rows


def build_review_candidates_rows(notices: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    review_notices = [
        x for x in notices
        if x.get("_label") in {"Direct", "Adjacent"}
    ]

    label_order = {"Direct": 0, "Adjacent": 1}
    review_notices.sort(
        key=lambda x: (
            label_order.get(str(x.get("_label", "")), 99),
            -int(x.get("_score", 0)),
            str(x.get("_record_id", "")),
        )
    )

    headers = [
        "review_item_key",
        "original_label",
        "final_decision",
        "action_status",
        "owner_note",
        "score",
        "source_kind",
        "source_type",
        "title",
        "pub_org",
        "close_dt",
        "query_keywords",
        "reasons",
        "bid_detail_url",
    ]

    rows: list[list[Any]] = []
    for item in review_notices:
        rows.append(
            [
                _to_text(item.get("_record_id")),
                _to_text(item.get("_label")),
                "",
                "new",
                "",
                item.get("_score", 0),
                _to_text(item.get("_source_kind")),
                _to_text(item.get("_source_type")),
                _to_text(item.get("_title")),
                _to_text(item.get("_pub_org")),
                _to_text(item.get("_close_dt")),
                _to_text(item.get("_query_keywords")),
                _to_text(item.get("_reasons")),
                _to_text(item.get("_bid_detail_url")),
            ]
        )

    return headers, rows


def build_attachments_rows(download_results: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "record_id",
        "seq",
        "file_name",
        "file_url",
        "local_path",
        "status",
        "error",
    ]

    rows: list[list[Any]] = []
    for item in download_results:
        rows.append(
            [
                _to_text(item.get("record_id")),
                _to_text(item.get("seq")),
                _to_text(item.get("file_name")),
                _to_text(item.get("file_url")),
                _to_text(item.get("local_path")),
                _to_text(item.get("status")),
                _to_text(item.get("error")),
            ]
        )

    return headers, rows


def build_summary_rows(
    manifest: dict[str, Any],
    raw_json_path: str,
    total_items: int,
    deduped_items: int,
    notices: list[dict[str, Any]],
    download_results: list[dict[str, Any]],
) -> list[list[Any]]:
    direct_count = sum(1 for x in notices if x.get("_label") == "Direct")
    adjacent_count = sum(1 for x in notices if x.get("_label") == "Adjacent")
    exclude_count = sum(1 for x in notices if x.get("_label") == "Exclude")

    downloaded_count = sum(1 for x in download_results if x.get("status") == "DOWNLOADED")
    failed_count = sum(1 for x in download_results if x.get("status") == "FAILED")

    enrich_summary = manifest.get("enrich_summary", {})

    rows = [
        ["generated_at", datetime.now().isoformat(timespec="seconds")],
        ["raw_json_path", raw_json_path],
        ["total_items", total_items],
        ["deduped_items", deduped_items],
        ["final_items", len(notices)],
        ["direct_count", direct_count],
        ["adjacent_count", adjacent_count],
        ["exclude_count", exclude_count],
        ["attachment_jobs", len(download_results)],
        ["attachment_downloaded", downloaded_count],
        ["attachment_failed", failed_count],
        ["enrich_attempted", enrich_summary.get("attempted", 0)],
        ["enrich_merged", enrich_summary.get("merged", 0)],
        ["enrich_not_found", enrich_summary.get("not_found", 0)],
        ["enrich_error", enrich_summary.get("error", 0)],
        ["range_start_dt", _to_text(manifest.get("range", {}).get("start_dt"))],
        ["range_end_dt", _to_text(manifest.get("range", {}).get("end_dt"))],
        ["attachments_scope", _to_text(manifest.get("attachments_scope"))],
        ["detail_url_policy", _to_text(manifest.get("detail_url_policy"))],
        ["classification_policy_version", _to_text(manifest.get("classification_policy_version"))],
        ["date_range_source", _to_text(manifest.get("date_range_source"))],
        ["lookback_days", _to_text(manifest.get("lookback_days"))],
    ]
    return rows


def export_run_to_excel(
    notices: list[dict[str, Any]],
    download_results: list[dict[str, Any]],
    manifest: dict[str, Any],
    raw_json_path: str,
    total_items: int,
    deduped_items: int,
) -> str:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    candidates_ws = wb.active
    candidates_ws.title = "candidates"

    candidate_headers, candidate_rows = build_candidates_rows(notices)
    _write_sheet(candidates_ws, candidate_headers, candidate_rows)
    _apply_label_fill(candidates_ws, "label")
    _apply_url_links(candidates_ws, ["bid_detail_url"])

    review_ws = wb.create_sheet("review_candidates")
    review_headers, review_rows = build_review_candidates_rows(notices)
    _write_sheet(review_ws, review_headers, review_rows)
    _apply_label_fill(review_ws, "original_label")
    _apply_url_links(review_ws, ["bid_detail_url"])
    _highlight_input_columns(review_ws, ["final_decision", "action_status", "owner_note"])
    _add_dropdown_validation(review_ws, "final_decision", FINAL_DECISION_OPTIONS)
    _add_dropdown_validation(review_ws, "action_status", ACTION_STATUS_OPTIONS)

    attachments_ws = wb.create_sheet("attachments")
    attachment_headers, attachment_rows = build_attachments_rows(download_results)
    _write_sheet(attachments_ws, attachment_headers, attachment_rows)
    _apply_url_links(attachments_ws, ["file_url"])

    summary_ws = wb.create_sheet("run_summary")
    summary_ws.append(["key", "value"])
    for row in build_summary_rows(
        manifest=manifest,
        raw_json_path=raw_json_path,
        total_items=total_items,
        deduped_items=deduped_items,
        notices=notices,
        download_results=download_results,
    ):
        summary_ws.append(row)

    _style_header(summary_ws)
    _style_body(summary_ws)
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 80

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"day2_candidates_{timestamp}.xlsx"
    wb.save(output_path)

    return str(output_path)
